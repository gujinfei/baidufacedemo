import urllib.request
import json
import datetime
import base64
import openpyxl
import result_enum as re
import const
import time
import os
from common_function import trave_all_pic_file

const.max_angle = 20

# client_id 为官网获取的AK， client_secret 为官网获取的SK
def prepare(client_id,client_secret):
    try:
        host = 'https://aip.baidubce.com/oauth/2.0/token?grant_type=client_credentials&client_id=' + \
               client_id + '&client_secret=' + client_secret
        headers = {"Content-Type": "application/json; charset=UTF-8"}
        req = urllib.request.Request(url=host, headers=headers)
        response = urllib.request.urlopen(req,timeout=2)
        content = response.read()
        if (content):
            content.decode('utf-8')
            jsonData = json.loads(content)
            # print(jsonData)
            return jsonData['access_token'],jsonData['expires_in']
    except Exception as e:  # 抛出超时异常
        print('a', str(e))

def faceDetect(accessToken, base64pic):
    try:
        url = "https://aip.baidubce.com/rest/2.0/face/v3/detect"
        access_token = accessToken
        url = url + "?access_token=" + access_token

        headers = {"Content-Type": "application/json; charset=UTF-8"}
        values = {}
        values['image'] = base64pic
        values['image_type'] = 'BASE64'
        values['face_field'] = 'age,beauty,expression,face_shape,gender,facetype,emotion,quality,glasses,eye_status,race'
        values['max_face_num'] = 1
        data = urllib.parse.urlencode(values).encode(encoding='UTF8')
        req = urllib.request.Request(url, data, headers)
        response = urllib.request.urlopen(req, timeout=2)
        content = response.read()
        if content:
            content.decode('utf-8')
            jsonData = json.loads(content)
            # print(jsonData)
            return jsonData
    except Exception as e:  # 抛出超时异常
        print('a', str(e))

def persistent(fileName, result):
    # workbook = openpyxl.Workbook(encoding='UTF-8')
    workbook = openpyxl.Workbook()
    # 获取当前活跃的worksheet,默认就是第一个worksheet
    worksheet = workbook.active
    worksheet.title = "baidu"

    for i in range(len(re.Project)):
        worksheet.cell(1, i + 1, re.Project[i])

    for i in range(len(result)):
        faceResult = result[i]
        worksheet.cell(i + 2, re.ResultEnum.NAME.value, faceResult['name'])
        worksheet.cell(i + 2, re.ResultEnum.PROBABILITY.value, faceResult['result']['face_list'][0]['face_probability'])
        worksheet.cell(i + 2, re.ResultEnum.GENDER.value, faceResult['result']['face_list'][0]['gender']['type'])
        worksheet.cell(i + 2, re.ResultEnum.GENDERRATE.value, faceResult['result']['face_list'][0]['gender']['probability'])
        worksheet.cell(i + 2, re.ResultEnum.AGE.value, faceResult['result']['face_list'][0]['age'])

        worksheet.cell(i + 2, re.ResultEnum.RACE.value, faceResult['result']['face_list'][0]['race']['type'])
        worksheet.cell(i + 2, re.ResultEnum.RACERATE.value, faceResult['result']['face_list'][0]['race']['probability'])
        worksheet.cell(i + 2, re.ResultEnum.EXPRESSION.value, faceResult['result']['face_list'][0]['expression']['type'])
        worksheet.cell(i + 2, re.ResultEnum.EXPRESSIONRATE.value, faceResult['result']['face_list'][0]['expression']['probability'])
        worksheet.cell(i + 2, re.ResultEnum.FACESHARP.value, faceResult['result']['face_list'][0]['face_shape']['type'])

        worksheet.cell(i + 2, re.ResultEnum.FACESHARPRATE.value,faceResult['result']['face_list'][0]['face_shape']['probability'])
        worksheet.cell(i + 2, re.ResultEnum.FACETYPE.value, faceResult['result']['face_list'][0]['face_type']['type'])
        worksheet.cell(i + 2, re.ResultEnum.FACETYPERATE.value,faceResult['result']['face_list'][0]['face_type']['probability'])
        worksheet.cell(i + 2, re.ResultEnum.FACEEMOTION.value, faceResult['result']['face_list'][0]['emotion']['type'])
        worksheet.cell(i + 2, re.ResultEnum.FACEEMOTIONRATE.value,faceResult['result']['face_list'][0]['emotion']['probability'])

        yaw = faceResult['result']['face_list'][0]['angle']['yaw']
        pitch = faceResult['result']['face_list'][0]['angle']['pitch']
        roll = faceResult['result']['face_list'][0]['angle']['roll']
        worksheet.cell(i + 2, re.ResultEnum.YAW.value, yaw)
        worksheet.cell(i + 2, re.ResultEnum.PITCH.value, pitch)
        worksheet.cell(i + 2, re.ResultEnum.ROLL.value, roll)
        angle = 1
        if abs(yaw) > const.max_angle or abs(pitch) > const.max_angle or abs(roll) > const.max_angle:
            angle = 0
        worksheet.cell(i + 2, re.ResultEnum.ANGLEOK.value, angle)
        worksheet.cell(i + 2, re.ResultEnum.LEFTEYE.value, faceResult['result']['face_list'][0]['quality']['occlusion']['left_eye'])

        worksheet.cell(i + 2, re.ResultEnum.RIGHTEYE.value, faceResult['result']['face_list'][0]['quality']['occlusion']['right_eye'])
        worksheet.cell(i + 2, re.ResultEnum.NOSE.value, faceResult['result']['face_list'][0]['quality']['occlusion']['nose'])
        worksheet.cell(i + 2, re.ResultEnum.MOUTH.value, faceResult['result']['face_list'][0]['quality']['occlusion']['mouth'])
        worksheet.cell(i + 2, re.ResultEnum.LEFTCHEEK.value, faceResult['result']['face_list'][0]['quality']['occlusion']['left_cheek'])
        worksheet.cell(i + 2, re.ResultEnum.RIGHTCHEEK.value, faceResult['result']['face_list'][0]['quality']['occlusion']['right_cheek'])

        worksheet.cell(i + 2, re.ResultEnum.CHIN.value, faceResult['result']['face_list'][0]['quality']['occlusion']['chin_contour'])
        shadow = 0
        worksheet.cell(i + 2, re.ResultEnum.SHAOW.value,shadow)
        worksheet.cell(i + 2, re.ResultEnum.BLUR.value, faceResult['result']['face_list'][0]['quality']['blur'])
        worksheet.cell(i + 2, re.ResultEnum.ILLUMINATION.value, faceResult['result']['face_list'][0]['quality']['illumination'])
        worksheet.cell(i + 2, re.ResultEnum.COMPLETENESS.value, faceResult['result']['face_list'][0]['quality']['completeness'])

        worksheet.cell(i + 2, re.ResultEnum.GLASSTYPE.value, faceResult['result']['face_list'][0]['glasses']['type'])
        worksheet.cell(i + 2, re.ResultEnum.GLASSRATE.value, faceResult['result']['face_list'][0]['glasses']['probability'])
        worksheet.cell(i + 2, re.ResultEnum.LEFTEYEOPEN.value, faceResult['result']['face_list'][0]['eye_status']['left_eye'])
        worksheet.cell(i + 2, re.ResultEnum.RIGHTEYEOPEN.value,faceResult['result']['face_list'][0]['eye_status']['right_eye'])

    workbook.save(filename=fileName)

if __name__ == '__main__':
    pic_dir = input("请输入图片所在根目录：")
    while not os.path.exists(pic_dir):
        pic_dir = input("目录不存在，请输入图片所在根目录：")
    result_file = input("请输入结果excel全路径文件名:")
    parent_path = os.path.dirname(result_file)
    while not os.path.exists(parent_path):
        result_file = input("目录不存在，请输入结果excel全路径文件名:")
        path = os.path.dirname(result_file)
    fault_file = input("请输入错误日志文件全路径文件名(txt):")
    path = os.path.dirname(fault_file)
    while not os.path.exists(path):
        fault_file = input("目录不存在，请输入错误日志文件全路径文件名(txt):")
        path = os.path.dirname(fault_file)

    start = datetime.datetime.now()
    toke_info = prepare('xqzaGjQe4xq7qcaDYXyk3edg', '4coHiG4ofcGCKTklVIGkPzYc0GIDDmSl')

    if toke_info is None:
        toke_info = prepare('xqzaGjQe4xq7qcaDYXyk3edg', '4coHiG4ofcGCKTklVIGkPzYc0GIDDmSl')

    if toke_info is None:
        print("ERROR:failed to get token info!!")
        sys.exit("Goodbye!");

    # allPicFiles = trave_all_pic_file("/Volumes/Untitled/3万张图片质量分类-190415/角度")
    allPicFiles = trave_all_pic_file(pic_dir)
    if len(allPicFiles) == 0:
        print("no valid pic file")

    resultList = []
    errorFileList = []
    for i in range(len(allPicFiles)):
        print("***** procee all count is %d, current is %d ******" % (len(allPicFiles), i+1))
        file = allPicFiles[i]
        # print("process file:%s" % file)
        with open(file, 'rb') as f:  # 以二进制读取图片
            data = f.read()
            encodestr = base64.b64encode(data)  # 得到 byte 编码的数据
            result = faceDetect(toke_info[0], encodestr)
            if result is None:
                count = 0
                while count < 6 and result is None:
                    count = count + 1
                    time.sleep(count)
                    result = faceDetect(toke_info[0], encodestr)
            if result is None:
                print("ERROR:net error! open or read faild")
                info = file + ' ERROR:net error! open or read faild\n'
                errorFileList.append(info)
            elif result['error_code'] == 0:
                result['name'] = file
                resultList.append(result)
            else:
                info = file + " ERROR:error code is " + str(result['error_code']) + '\n'
                errorFileList.append(info)

    # persistent('/Users/gujinfei/baidu.xlsx', resultList)
    # f = open('/Users/gujinfei/baidu.txt', 'a')
    persistent(result_file, resultList)
    f = open(fault_file, 'w')
    f.writelines(errorFileList)
    f.close()
    print("all success!![success:%d, failed:%d]"%(len(resultList), len(errorFileList)))

