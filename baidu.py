# -*- coding: utf-8 -*-
import json
import openpyxl
import const
import urllib.request
from datetime import datetime
import result_enum as re

const.CLIENTKEY = 'xqzaGjQe4xq7qcaDYXyk3edg'
const.SECRETKEY = '4coHiG4ofcGCKTklVIGkPzYc0GIDDmSl'

ACCESS_TOKEN = ''
TOKEN_EXPIRE_SECOND = 0
START_TIME = datetime.now()
const.max_angle = 20

def initialize():
    try:
        host = 'https://aip.baidubce.com/oauth/2.0/token?grant_type=client_credentials&client_id=' + \
               const.CLIENTKEY + '&client_secret=' + const.SECRETKEY
        headers = {"Content-Type": "application/json; charset=UTF-8"}
        req = urllib.request.Request(url=host, headers=headers)
        response = urllib.request.urlopen(req,timeout=2)
        content = response.read()
        if (content):
            content.decode('utf-8')
            global START_TIME
            global ACCESS_TOKEN
            global TOKEN_EXPIRE_SECOND
            jsonData = json.loads(content)
            ACCESS_TOKEN = jsonData['access_token']
            TOKEN_EXPIRE_SECOND = jsonData['expires_in']
            START_TIME = datetime.now()
            return True
        return False
    except Exception as e:  # 抛出超时异常
        print(str(e))
        return False

def faceDetect(base64pic):
    now = datetime.now()
    if (now-START_TIME).seconds > TOKEN_EXPIRE_SECOND + 100:
        loop = 0
        while loop < 5 and not initialize():
            loop = loop + 1;
        if loop == 5:
            print("ERROR:baidu get token failed")
            return None
    try:
        url = "https://aip.baidubce.com/rest/2.0/face/v3/detect"
        access_token = ACCESS_TOKEN
        url = url + "?access_token=" + access_token

        headers = {"Content-Type": "application/json; charset=UTF-8"}
        values = {}
        values['image'] = base64pic
        values['image_type'] = 'BASE64'
        values['face_field'] = 'age,beauty,expression,face_shape,gender,facetype,emotion,quality,glasses,eye_status,race'
        values['max_face_num'] = 1
        data = urllib.parse.urlencode(values).encode(encoding='UTF8')
        req = urllib.request.Request(url, data, headers)
        response = urllib.request.urlopen(req, timeout=5)
        content = response.read()
        if content:
            content.decode('utf-8')
            jsonData = json.loads(content)
            # print(jsonData)
            return jsonData
    except Exception as e:  # 抛出超时异常
        print('a', str(e))

def persistent(xlxsFile, txtFile, result, errorFileList):
    # workbook = openpyxl.Workbook(encoding='UTF-8')
    workbook = openpyxl.Workbook()
    # 获取当前活跃的worksheet,默认就是第一个worksheet
    worksheet = workbook.active
    worksheet.title = "baidu"

    for i in range(len(re.Project)):
        worksheet.cell(1, i + 1, re.Project[i])

    for i in range(len(result)):
        faceResult = result[i]
        worksheet.cell(i + 2, re.ResultEnum.NAME.value, faceResult['name'])
        worksheet.cell(i + 2, re.ResultEnum.PROBABILITY.value, faceResult['result']['face_list'][0]['face_probability'])
        worksheet.cell(i + 2, re.ResultEnum.GENDER.value, faceResult['result']['face_list'][0]['gender']['type'])
        worksheet.cell(i + 2, re.ResultEnum.GENDERRATE.value, faceResult['result']['face_list'][0]['gender']['probability'])
        worksheet.cell(i + 2, re.ResultEnum.AGE.value, faceResult['result']['face_list'][0]['age'])

        worksheet.cell(i + 2, re.ResultEnum.RACE.value, faceResult['result']['face_list'][0]['race']['type'])
        worksheet.cell(i + 2, re.ResultEnum.RACERATE.value, faceResult['result']['face_list'][0]['race']['probability'])
        worksheet.cell(i + 2, re.ResultEnum.EXPRESSION.value, faceResult['result']['face_list'][0]['expression']['type'])
        worksheet.cell(i + 2, re.ResultEnum.EXPRESSIONRATE.value, faceResult['result']['face_list'][0]['expression']['probability'])
        worksheet.cell(i + 2, re.ResultEnum.FACESHARP.value, faceResult['result']['face_list'][0]['face_shape']['type'])

        worksheet.cell(i + 2, re.ResultEnum.FACESHARPRATE.value,faceResult['result']['face_list'][0]['face_shape']['probability'])
        worksheet.cell(i + 2, re.ResultEnum.FACETYPE.value, faceResult['result']['face_list'][0]['face_type']['type'])
        worksheet.cell(i + 2, re.ResultEnum.FACETYPERATE.value,faceResult['result']['face_list'][0]['face_type']['probability'])
        worksheet.cell(i + 2, re.ResultEnum.FACEEMOTION.value, faceResult['result']['face_list'][0]['emotion']['type'])
        worksheet.cell(i + 2, re.ResultEnum.FACEEMOTIONRATE.value,faceResult['result']['face_list'][0]['emotion']['probability'])

        yaw = faceResult['result']['face_list'][0]['angle']['yaw']
        pitch = faceResult['result']['face_list'][0]['angle']['pitch']
        roll = faceResult['result']['face_list'][0]['angle']['roll']
        worksheet.cell(i + 2, re.ResultEnum.YAW.value, yaw)
        worksheet.cell(i + 2, re.ResultEnum.PITCH.value, pitch)
        worksheet.cell(i + 2, re.ResultEnum.ROLL.value, roll)
        angle = 1
        if abs(yaw) > const.max_angle or abs(pitch) > const.max_angle or abs(roll) > const.max_angle:
            angle = 0
        worksheet.cell(i + 2, re.ResultEnum.ANGLEOK.value, angle)
        worksheet.cell(i + 2, re.ResultEnum.LEFTEYE.value, faceResult['result']['face_list'][0]['quality']['occlusion']['left_eye'])

        worksheet.cell(i + 2, re.ResultEnum.RIGHTEYE.value, faceResult['result']['face_list'][0]['quality']['occlusion']['right_eye'])
        worksheet.cell(i + 2, re.ResultEnum.NOSE.value, faceResult['result']['face_list'][0]['quality']['occlusion']['nose'])
        worksheet.cell(i + 2, re.ResultEnum.MOUTH.value, faceResult['result']['face_list'][0]['quality']['occlusion']['mouth'])
        worksheet.cell(i + 2, re.ResultEnum.LEFTCHEEK.value, faceResult['result']['face_list'][0]['quality']['occlusion']['left_cheek'])
        worksheet.cell(i + 2, re.ResultEnum.RIGHTCHEEK.value, faceResult['result']['face_list'][0]['quality']['occlusion']['right_cheek'])

        worksheet.cell(i + 2, re.ResultEnum.CHIN.value, faceResult['result']['face_list'][0]['quality']['occlusion']['chin_contour'])
        shadow = 0
        worksheet.cell(i + 2, re.ResultEnum.SHAOW.value,shadow)
        worksheet.cell(i + 2, re.ResultEnum.BLUR.value, faceResult['result']['face_list'][0]['quality']['blur'])
        worksheet.cell(i + 2, re.ResultEnum.ILLUMINATION.value, faceResult['result']['face_list'][0]['quality']['illumination'])
        worksheet.cell(i + 2, re.ResultEnum.COMPLETENESS.value, faceResult['result']['face_list'][0]['quality']['completeness'])

        worksheet.cell(i + 2, re.ResultEnum.GLASSTYPE.value, faceResult['result']['face_list'][0]['glasses']['type'])
        worksheet.cell(i + 2, re.ResultEnum.GLASSRATE.value, faceResult['result']['face_list'][0]['glasses']['probability'])
        worksheet.cell(i + 2, re.ResultEnum.LEFTEYEOPEN.value, faceResult['result']['face_list'][0]['eye_status']['left_eye'])
        worksheet.cell(i + 2, re.ResultEnum.RIGHTEYEOPEN.value,faceResult['result']['face_list'][0]['eye_status']['right_eye'])

    workbook.save(filename=xlxsFile)
    f = open(txtFile, 'w')
    f.writelines(errorFileList)
    f.close()